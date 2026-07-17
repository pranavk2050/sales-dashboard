"""Generates ./data/sample.xlsx matching the real "Opportunity and lead tracker
sheet_Digital.xlsx" schema (column layouts and quirks confirmed against that workbook,
not just the original written spec). Includes intentional edge cases:
- a blank Opp/Lead No. row and a "YTR" placeholder row (both must get a synthetic key,
  not be dropped or collide)
- a duplicate real Opp/Lead No. (must be disambiguated, not silently overwritten)
- a #REF! value (must become null)
- an Inter BU project-less BU/team status row (must get a placeholder project)
- a Lost-opportunities row with the real sheet's column-drift quirk (extra 0/1 flag
  shifting quarter/reason/date/names one column right)

Run: py backend/scripts/generate_sample.py
"""
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BOLD = Font(bold=True)

MAIN_HEADERS = [
    "SL No.", "Opp/Lead No.", "Customer", "Enquiry description", "Tentative value (INR in Cr)",
    "Opportunity", "Progress", "Expected", "Present status", "Timeline",
    "Delivery team involved", "Sales team involved",
]

# (sl_no, opp_no, customer, enquiry, value, multiplier, progress, expected, status, timeline, delivery, sales)
MAIN_ROWS = [
    (1, "O-ATBU-012026-0001", "Tata Steel-TSK", "Blast furnace automation upgrade", 4, "5X", 1, "Q3",
     "Price submitted, awaiting client's feedback", date(2026, 6, 20), "Team A", "R. Sharma"),
    (2, "O-ATBU-022026-0015", "IOCL", "Refinery instrumentation retrofit", 17, "10X", 0, "Q3",
     "Price submitted, awaiting client's feedback", date(2026, 7, 10), "Team A", "R. Sharma"),
    (3, "O-ATBU-032026-0002", "BPCL Mumbai refinery", "Flow meter calibration project", 0.5, "5X", 1, "Q2",
     "PO awaited", date(2026, 7, 15), "Team B", "P. Iyer"),
    (4, "O-ATBU-032026-0003", "Indorama", "DCS upgrade feasibility study", 0.25, "5X", 0, "Q1",
     "Lost to competitor - reconsidering", date(2026, 7, 5), "Team C", "S. Rao"),
    (5, "O-ATBU-042026-0004", "KNPC", "Control valve replacement", 8, "7X", 1, "Q3",
     "Site survey completed", date(2026, 7, 18), "Team A", "R. Sharma"),
    (6, "O-ATBU-042026-0005", "KBL", "Pump station SCADA integration", "#REF!", "10X", 0, "Q4",
     "Awaiting client PO", date(2026, 7, 25), "Team D", "M. Nair"),
    (7, "O-ATBU-052026-0006", "Tata Steel-TSK", "Coke oven gas analyzer", 3, "5X", 1, "Q3",
     "Order under negotiation", date(2026, 6, 1), "Team B", "P. Iyer"),
    (8, "O-ATBU-052026-0007", "IOCL", "Tank farm level monitoring", 6, "7X", 1, "Q2",
     "Client site visit scheduled", date(2026, 7, 17), "Team A", "R. Sharma"),
    (9, "O-ATBU-062026-0008", "BPCL Mumbai refinery", "Emergency shutdown system audit", 20, "15X", 0, "Q4",
     "Proposal under review", date(2026, 6, 30), "Team C", "S. Rao"),
    (10, "O-ATBU-062026-0009", "Indorama", "Compressor vibration monitoring", 1.5, "5X", 1, "Q3",
     "Won - kickoff pending", date(2026, 7, 14), "Team D", "M. Nair"),
    (11, "O-ATBU-072026-0010", "KNPC", "Fire & gas system upgrade", 9, "10X", 0, "Q1",
     "RFQ received", date(2026, 7, 20), "Team B", "P. Iyer"),
    (12, "O-ATBU-072026-0011", "KBL", "Motor control center replacement", 2, "5X", 1, "Q2",
     "Site walk-down done", date(2026, 7, 22), "Team A", "R. Sharma"),
]
# Row 9 (30.06.2026) is exactly 16 days before 16.07.2026 "today" -> designated Phase 3 alert row.
# Row 3 (15.07.2026) and row 8 (17.07.2026) fall in the current week -> Phase 3 "this week" panel.
# Row 2 keeps timeline=10.07.2026 matching the stakeholder's verbatim Phase 2 example
# (extend to 16.07.2026 later to test the change log).

# Edge cases seen in the real workbook: appended after MAIN_ROWS.
MAIN_EXTRA_ROWS = [
    # Blank Opp/Lead No. but real lead data -> must get a synthetic "LEAD-MAIN-<row>" key.
    (13, None, "BWSSB", "Digital twin for water board", None, "20X", 0, None,
     "TCP submitted, need to follow up", date(2026, 7, 20), None, "Saikat"),
    # "YTR" placeholder ("Yet To Raise") -> must also get a synthetic key, not the literal string.
    (14, "YTR", "Vedanta Lanjigarh", "BMS solutions enquiry", None, "15X", 0, None,
     "Meeting held, customer will send enquiry", date(2026, 7, 28), None, "Saikat"),
    # Duplicate of row 1's real Opp/Lead No. -> must be disambiguated (#2 suffix), not overwrite row 1.
    (15, "O-ATBU-012026-0001", "KBL", "Duplicate-ID regression test row", 3, "7X", 1, "Q3",
     "Tests duplicate key disambiguation", date(2026, 7, 19), "Team C", "M. Nair"),
    # Fully blank row (all None) -> must be skipped entirely, unlike the rows above.
    (16, None, None, None, None, None, None, None, None, None, None, None),
]

LOST_HEADERS = [
    "SL NO", "Opportunity no.", "Customer", "Services", "Value", "Status", "qaurter",
    "Lost reason", "Date lost", "Team member 1", "Team member 2", "Notes",
]

# Real layout: header row 1, data from row 2. (sl_no, opp_no, customer, services, value,
# multiplier, quarter, lost_reason, date_lost, team1, team2, notes)
LOST_ROWS = [
    (1, "O-ATBU-012026-0012", "Adani Power", "Boiler control retrofit", 5, "5X", "Q1",
     "Lost - client chose competitor", date(2026, 5, 1), "R. Sharma", "Team A", "Follow up next cycle"),
    (2, "O-ATBU-022026-0013", "Chambal Fertilizer", "Ammonia plant instrumentation", 3, "7X", "Q2",
     "Lost - budget cut", date(2026, 4, 15), "S. Rao", "Team B", None),
    # "YTR" placeholder in the Lost tab too -> synthetic key, no collision with other YTR rows.
    (3, "YTR", "DCM Shriram", "Effluent treatment automation", 1, "5X", "Q1",
     "Lost - project shelved", date(2026, 3, 20), "M. Nair", "Team C", None),
]

# The real sheet drifts one column right from roughly row 12 on: an extra 0/1 flag lands
# between the multiplier and the quarter. (sl_no, opp_no, customer, services, value,
# multiplier, stray_flag, quarter, lost_reason, date_lost, team1, team2, notes)
LOST_SHIFTED_ROW = (
    4, "O-ATBU-032026-0099", "Simulated Shifted Co", "Tests column-drift heuristic", 2, "10X",
    0, "Q4", "Lost - shifted column test", date(2026, 6, 10), "Team D", "M. Nair", "drift test",
)

INTERBU_HEADERS = [
    "SL No.", "BUs", "Team members involved", "Meeting status", "Present Status", "Projects",
    "Service Lines", "Status", "Stage", "Responsibility", "Target date",
    "Tentative value(INR in Cr)", "Name", "Designation",
]

# group_header rows use only the BUs column; project rows carry sl_no. Real grouping labels
# are "Domestic" and "ISMG" (not "International"). A couple of rows deliberately have no
# Project (a BU/team status update) to exercise the "(General - <team>)" placeholder path.
INTERBU_ROWS = [
    {"group": "Domestic"},
    {"sl_no": 1, "bu": "HCBU", "team": "Satyajit and team", "meeting_status": "Meeting done",
     "present_status": "Meeting done, shared contact details", "project": "Chambal Fertilizer",
     "service_lines": "Automation", "status": "Customer contacted", "stage": "20X",
     "responsible": "Satyajit", "target_date": date(2026, 7, 25), "value": 6, "name": "Cadence Jul",
     "designation": "BU Head"},
    {"sl_no": None, "bu": None, "team": None, "meeting_status": None, "present_status": None,
     "project": "Matix Fertilizer", "service_lines": "Automation", "status": "Proposal shared",
     "stage": "20X", "responsible": "Satyajit", "target_date": date(2026, 7, 28), "value": 4,
     "name": "Cadence Jul", "designation": None},
    {"sl_no": 2, "bu": "PBU", "team": "Ganga", "meeting_status": "Meeting done",
     "present_status": "Need to align for Adani Power", "project": "Adani Power",
     "service_lines": "Controls", "status": "Meeting to be scheduled", "stage": "10X",
     "responsible": "Ganga", "target_date": date(2026, 7, 30), "value": 10, "name": "Cadence Jul",
     "designation": "BU Head"},
    # Project-less BU status update -> "(General - Sathish Rao)" placeholder.
    {"sl_no": 3, "bu": "PMC", "team": "Sathish Rao", "meeting_status": "Meeting pending",
     "present_status": "4D Synchro to be incorporated in deck", "project": None,
     "service_lines": None, "status": None, "stage": None, "responsible": None,
     "target_date": None, "value": None, "name": None, "designation": "Moved to main sheet"},
    {"group": "ISMG"},
    {"sl_no": 1, "bu": "Africa", "team": "Gaurav", "meeting_status": "Meeting done",
     "present_status": "PPT to be sent, monthly meeting to be fixed", "project": "Nothing",
     "service_lines": None, "status": None, "stage": None, "responsible": None,
     "target_date": None, "value": None, "name": None, "designation": None},
    {"sl_no": 2, "bu": "Europe", "team": "PK Suraj", "meeting_status": "Meeting done",
     "present_status": None, "project": "Stegra", "service_lines": "Digital twin",
     "status": None, "stage": None, "responsible": None, "target_date": None, "value": None,
     "name": None, "designation": None},
    # Second project-less row under the SAME BU label ("Europe") but a different team member ->
    # tests that the placeholder project doesn't collide with the first Europe entry above.
    {"sl_no": 3, "bu": "Europe", "team": "Rajesh Das", "meeting_status": "Meeting pending",
     "present_status": None, "project": None, "service_lines": None, "status": None,
     "stage": None, "responsible": None, "target_date": None, "value": None, "name": None,
     "designation": None},
]


def write_main_sheet(ws, rows, extra_rows):
    ws.append([])
    ws.append([])
    ws.append(MAIN_HEADERS)
    for cell in ws[3]:
        cell.font = BOLD
    for row in rows:
        ws.append(row)
    for row in extra_rows:
        ws.append(row)
    # highlight: overdue timeline red, healthy status green (cosmetic only, parser must ignore)
    for r in range(4, 4 + len(rows)):
        status_cell = ws.cell(row=r, column=9)
        if status_cell.value and "won" in str(status_cell.value).lower():
            status_cell.fill = GREEN
        timeline_cell = ws.cell(row=r, column=10)
        if isinstance(timeline_cell.value, date) and timeline_cell.value < date(2026, 7, 16):
            timeline_cell.fill = RED


def write_lost_sheet(ws):
    ws.append(LOST_HEADERS)
    for cell in ws[1]:
        cell.font = BOLD
    for row in LOST_ROWS:
        ws.append(row)
    # Write the shifted row's cells directly so the stray flag lands in its own column.
    sl_no, opp_no, customer, services, value, multiplier, stray_flag, quarter, lost_reason, date_lost, team1, team2, notes = LOST_SHIFTED_ROW
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=sl_no)
    ws.cell(row=r, column=2, value=opp_no)
    ws.cell(row=r, column=3, value=customer)
    ws.cell(row=r, column=4, value=services)
    ws.cell(row=r, column=5, value=value)
    ws.cell(row=r, column=6, value=multiplier)
    ws.cell(row=r, column=7, value=stray_flag)
    ws.cell(row=r, column=8, value=quarter)
    ws.cell(row=r, column=9, value=lost_reason)
    ws.cell(row=r, column=10, value=date_lost)
    ws.cell(row=r, column=11, value=team1)
    ws.cell(row=r, column=12, value=team2)
    ws.cell(row=r, column=13, value=notes)


def write_interbu_sheet(ws):
    ws.append([])
    ws.append([])
    ws.append(INTERBU_HEADERS)
    for cell in ws[3]:
        cell.font = BOLD
    for entry in INTERBU_ROWS:
        r = ws.max_row + 1
        if "group" in entry:
            ws.cell(row=r, column=2, value=entry["group"])
            continue
        ws.cell(row=r, column=1, value=entry["sl_no"])
        ws.cell(row=r, column=2, value=entry["bu"])
        ws.cell(row=r, column=3, value=entry["team"])
        ws.cell(row=r, column=4, value=entry["meeting_status"])
        ws.cell(row=r, column=5, value=entry["present_status"])
        ws.cell(row=r, column=6, value=entry["project"])
        ws.cell(row=r, column=7, value=entry["service_lines"])
        ws.cell(row=r, column=8, value=entry["status"])
        ws.cell(row=r, column=9, value=entry["stage"])
        ws.cell(row=r, column=10, value=entry["responsible"])
        ws.cell(row=r, column=11, value=entry["target_date"])
        ws.cell(row=r, column=12, value=entry["value"])
        ws.cell(row=r, column=13, value=entry["name"])
        ws.cell(row=r, column=14, value=entry["designation"])
        if entry["meeting_status"] == "Meeting done":
            ws.cell(row=r, column=4).fill = GREEN


def build_workbook() -> Workbook:
    wb = Workbook()

    main_ws = wb.active
    main_ws.title = "Main sheet"
    write_main_sheet(main_ws, MAIN_ROWS, MAIN_EXTRA_ROWS)

    sheet1 = wb.create_sheet("Sheet1")
    sheet1.append(["(unused in v1)"])

    lt_ws = wb.create_sheet("Long term opportunities")
    lt_ws.append(["(unused in v1)"])

    interbu_ws = wb.create_sheet("Inter BU")
    write_interbu_sheet(interbu_ws)

    lost_ws = wb.create_sheet("Lost opportunities")
    write_lost_sheet(lost_ws)

    qtr_ws = wb.create_sheet("Qtr basis")
    qtr_ws.append(["(unused in v1)"])

    cross_bu_ws = wb.create_sheet("Cross BU opportunities")
    cross_bu_ws.append(["(unused in v1)"])

    return wb


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    out_path = DATA_DIR / "sample.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
